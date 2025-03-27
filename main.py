import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Количество генераторов и моментов времени
num_generators = 16
num_time_slots = 20  # Количество пар моментов времени

# Генерация случайных требований на включение генераторов в каждый момент времени
# time_requirements = np.random.randint(0, num_generators + 1, size=num_time_slots).tolist()
time_requirements = [3, 4, 16, 14, 10, 6, 0, 3, 1, 1, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

# Проверка, что количество требований соответствует количеству временных слотов
assert len(time_requirements) == num_time_slots, "Количество требований должно совпадать с количеством временных слотов"

# Инициализация таблицы (пустое поле для "-")
table = np.full((num_generators, num_time_slots * 2), "", dtype=object)

# Список для подсчета включений каждого генератора
generator_usage = [0] * num_generators

# Алгоритм распределения включений

for t in range(num_time_slots):
    # Сортируем генераторы по количеству их включений
    sorted_generators = sorted(range(num_generators), key=lambda x: (generator_usage[x], x))
    
    # Включаем необходимое количество генераторов для текущего момента времени "+"
    for g in sorted_generators[:time_requirements[t]]:
        table[g][t * 2] = "+"  # "+" момент времени
        table[g][t * 2 + 1] = "-"  # "-" момент времени
        generator_usage[g] += 1

# for t in range(num_time_slots):
#     # Сортируем генераторы по количеству их включений, а при равенстве - по убыванию индекса
#     sorted_generators = sorted(range(num_generators), key=lambda x: (generator_usage[x], -x))
    
#     # Включаем необходимое количество генераторов для текущего момента времени "+"
#     for g in sorted_generators[:time_requirements[t]]:
#         table[g][t * 2] = "+"  # "+" момент времени
#         table[g][t * 2 + 1] = "-"  # "-" момент времени
#         generator_usage[g] += 1

# Создание DataFrame для удобного просмотра
df = pd.DataFrame(table, index=[f"Generator {i+1}" for i in range(num_generators)],
                  columns=[f"Time {j+1}" for j in range(num_time_slots * 2)])

# Добавление столбца с суммой плюсов
df["Total +"] = (df == "+").sum(axis=1)

# Добавление пустого столбца для удобства чтения
df[" "] = ""  # Пустой столбец

# Добавление строки с количеством включенных генераторов под таблицей
time_requirements_row = []
for i in range(num_time_slots):
    time_requirements_row.extend([time_requirements[i], ""])  # Значение под "+" и пустое под "-"

# Дополнение строки пустыми значениями для остальных столбцов
time_requirements_row += [""] * (len(df.columns) - len(time_requirements_row))
df.loc["Total Generators On"] = time_requirements_row

# Сохранение таблицы в Excel
output_file = "generator_schedule2.xlsx"
df.to_excel(output_file, index=True)

# Открытие файла Excel для изменения стилей
wb = load_workbook(output_file)
ws = wb.active

# Определение стиля для светло-зеленого цвета
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Применение стиля к ячейкам с "+" и "-"
for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=2, max_col=ws.max_column - 2):
    for cell in row:
        if cell.value == "+" or cell.value == "-":
            cell.fill = green_fill

# Применение жирного шрифта к строке с итогами
from openpyxl.styles import Font
bold_font = Font(bold=True)
for cell in ws[ws.max_row]:
    cell.font = bold_font

# Сохранение изменений в Excel
wb.save(output_file)
print(f"Таблица сохранена в файл {output_file} с окрашенными ячейками и итогами.")